from __future__ import annotations

import json
import re
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.services.path_service import DATASETS_DIR, REPORTS_DIR, RUNS_DIR

SUPPORTED_SUFFIXES = {".xlsx", ".csv", ".tsv", ".dta"}
DATASET_LIST_CACHE_TTL_SECONDS = 60.0
PROFILE_SAMPLE_ROW_LIMIT = 5000
LARGE_DATASET_ROW_THRESHOLD = 10000
_DATASET_LIST_CACHE_LOCK = threading.Lock()
_DATASET_LIST_CACHE: dict[str, Any] = {
    "expires_at": 0.0,
    "signature": (),
    "items": [],
}
NULLISH_TOKENS = {
    "",
    "-",
    "--",
    "---",
    "—",
    "——",
    "na",
    "n/a",
    "nan",
    "none",
    "null",
    "暂无",
    "无",
}
NUMERIC_TEXT_PATTERN = re.compile(
    r"^\(?\s*[-+]?\d[\d,\s]*(?:\.\d+)?\s*\)?\s*(?:%|k|m|b|K|M|B|万|亿)?$"
)
HEADER_SCAN_LIMIT = 25
HEADER_LOOKAHEAD_ROWS = 8
MAX_HEADER_DEPTH = 3
METADATA_CELL_PATTERN = re.compile(r"[:：]\s*$")


def ensure_storage() -> None:
    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    RUNS_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def slugify(value: str) -> str:
    raw = value.strip() or "sheet"
    safe = re.sub(r"[^a-zA-Z0-9_-]+", "-", raw).strip("-").lower()
    digest = uuid.uuid5(uuid.NAMESPACE_DNS, raw).hex[:8]
    return f"{safe}-{digest}" if safe else f"sheet-{digest}"


def clean_scalar(value: Any) -> Any:
    if value is None:
        return None
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, Path):
        return str(value)
    return value


def clean_records(frame: pd.DataFrame, limit: int = 8) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    serialisable = frame.head(limit).copy()
    for column in serialisable.columns:
        if pd.api.types.is_datetime64_any_dtype(serialisable[column]):
            serialisable[column] = serialisable[column].dt.strftime("%Y-%m-%d %H:%M:%S")
    serialisable = serialisable.where(pd.notnull(serialisable), None)
    return [
        {str(key): clean_scalar(value) for key, value in row.items()}
        for row in serialisable.to_dict(orient="records")
    ]


def _profile_sample_frame(frame: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    if len(frame) <= PROFILE_SAMPLE_ROW_LIMIT:
        return frame, "full_dataset"

    positions = np.linspace(0, len(frame) - 1, PROFILE_SAMPLE_ROW_LIMIT, dtype=int)
    sampled = frame.iloc[pd.Index(positions).drop_duplicates()].copy()
    return sampled, "deterministic_even_sample"


def save_metadata(dataset_dir: Path, payload: dict[str, Any]) -> None:
    (dataset_dir / "metadata.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    from app.services.dataset_catalog_service import invalidate_dataset_list_cache

    invalidate_dataset_list_cache()
    with _DATASET_LIST_CACHE_LOCK:
        _DATASET_LIST_CACHE.update({"expires_at": 0.0, "signature": (), "items": []})


def _compact_column_summary(summary: dict[str, Any]) -> dict[str, Any]:
    return {
        "name": str(summary.get("name") or ""),
        "dtype": str(summary.get("dtype") or ""),
        "missing_count": int(summary.get("missing_count") or 0),
        "missing_ratio": float(summary.get("missing_ratio") or 0.0),
        "unique_count": int(summary.get("unique_count") or 0),
        "sample_values": [],
        "stats": {},
    }


def _dataset_list_item(metadata: dict[str, Any]) -> dict[str, Any]:
    """Return the lightweight shape used by dataset pickers.

    Full sample rows, chart bundles, and detailed stats remain available from
    GET /api/datasets/{dataset_id}; sending them for every row made the list
    endpoint several megabytes and slow to parse in the browser.
    """
    item = {
        "dataset_id": str(metadata.get("dataset_id") or ""),
        "name": str(metadata.get("name") or ""),
        "filename": str(metadata.get("filename") or ""),
        "active_sheet": str(metadata.get("active_sheet") or ""),
        "sheets": [
            {
                "name": str(sheet.get("name") or ""),
                "storage_file": str(sheet.get("storage_file") or ""),
                "rows": int(sheet.get("rows") or 0),
                "columns": int(sheet.get("columns") or 0),
            }
            for sheet in metadata.get("sheets", [])
            if isinstance(sheet, dict)
        ],
        "row_count": int(metadata.get("row_count") or 0),
        "column_count": int(metadata.get("column_count") or 0),
        "numeric_columns": [str(column) for column in metadata.get("numeric_columns", [])],
        "categorical_columns": [str(column) for column in metadata.get("categorical_columns", [])],
        "datetime_columns": [str(column) for column in metadata.get("datetime_columns", [])],
        "missing_cells": int(metadata.get("missing_cells") or 0),
        "sample_rows": [],
        "chart_bundle": {},
        "column_summaries": [
            _compact_column_summary(summary)
            for summary in metadata.get("column_summaries", [])
            if isinstance(summary, dict)
        ],
        "last_updated": str(metadata.get("last_updated") or ""),
    }
    if metadata.get("public_source"):
        item["public_source"] = str(metadata.get("public_source") or "")
    return item


def _worksheet_value_matrix(worksheet: Any) -> list[list[Any]]:
    matrix = [
        [cell.value for cell in row]
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        )
    ]
    if not matrix:
        return []

    for merged_range in worksheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        anchor = matrix[min_row - 1][min_col - 1]
        for row_index in range(min_row - 1, max_row):
            for col_index in range(min_col - 1, max_col):
                if matrix[row_index][col_index] is None:
                    matrix[row_index][col_index] = anchor

    return matrix


def _sheet_merge_metadata(worksheet: Any) -> dict[str, Any]:
    ranges = [str(item) for item in worksheet.merged_cells.ranges]
    return {
        "merged_cell_count": len(ranges),
        "merged_ranges_preview": ranges[:20],
    }


def _load_excel_workbook_sheets(raw_path: Path) -> list[tuple[str, pd.DataFrame, dict[str, Any]]]:
    workbook = load_workbook(raw_path, read_only=False, data_only=True)
    sheets: list[tuple[str, pd.DataFrame, dict[str, Any]]] = []
    for worksheet in workbook.worksheets:
        matrix = _worksheet_value_matrix(worksheet)
        frame = pd.DataFrame(matrix)
        sheets.append((worksheet.title, frame, _sheet_merge_metadata(worksheet)))
    return sheets


def _normalise_sheet_frame(frame: pd.DataFrame) -> pd.DataFrame:
    normalised = frame.copy()
    normalised = normalised.replace(r"^\s*$", np.nan, regex=True)
    normalised = normalised.dropna(how="all").dropna(axis=1, how="all")
    normalised = normalised.reset_index(drop=True)
    return normalised


def _normalise_object_cell(value: Any) -> Any:
    if value is None or pd.isna(value):
        return np.nan
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in NULLISH_TOKENS:
            return np.nan
        return stripped
    return value


def _parse_numeric_candidate(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and np.isnan(value)) or pd.isna(value):
        return None
    if isinstance(value, (bool, np.bool_)):
        return float(value)
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    if isinstance(value, (pd.Timestamp, datetime)):
        return None
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None

    multiplier = 1.0
    if text.endswith("%"):
        multiplier = 0.01
        text = text[:-1].strip()
    elif text.endswith(("k", "K")):
        multiplier = 1_000.0
        text = text[:-1].strip()
    elif text.endswith(("m", "M")):
        multiplier = 1_000_000.0
        text = text[:-1].strip()
    elif text.endswith(("b", "B")):
        multiplier = 1_000_000_000.0
        text = text[:-1].strip()
    elif text.endswith("万"):
        multiplier = 10_000.0
        text = text[:-1].strip()
    elif text.endswith("亿"):
        multiplier = 100_000_000.0
        text = text[:-1].strip()

    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1].strip()

    text = text.replace(",", "").replace(" ", "")
    if not NUMERIC_TEXT_PATTERN.match(f"{'-' if negative else ''}{text}"):
        try:
            return float(text) * multiplier * (-1.0 if negative else 1.0)
        except ValueError:
            return None

    try:
        return float(text) * multiplier * (-1.0 if negative else 1.0)
    except ValueError:
        return None


def _looks_like_date_value(value: Any) -> bool:
    if isinstance(value, (pd.Timestamp, datetime)):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    return bool(
        re.search(r"\d{4}[-/年]\d{1,2}[-/月]\d{1,2}", text)
        or re.search(r"\d{1,2}:\d{2}", text)
    )


def _coerce_object_series(series: pd.Series) -> pd.Series:
    if not pd.api.types.is_object_dtype(series) and not pd.api.types.is_string_dtype(series):
        return series

    cleaned = series.map(_normalise_object_cell)
    observed = cleaned.dropna()
    if observed.empty:
        return cleaned

    numeric_guess = observed.map(_parse_numeric_candidate)
    numeric_ratio = float(numeric_guess.notna().mean()) if len(observed) else 0.0
    if numeric_ratio >= 0.85 and numeric_guess.notna().sum() >= 2:
        converted = cleaned.map(_parse_numeric_candidate)
        numeric = pd.to_numeric(converted, errors="coerce")
        non_null = numeric.dropna()
        if not non_null.empty:
            numeric_values = non_null.to_numpy(dtype=float)
            integer_like = bool(np.isclose(numeric_values, np.round(numeric_values), atol=1e-12).all())
            if integer_like:
                rounded = pd.Series(
                    np.where(numeric.notna(), np.round(numeric.astype(float)), np.nan),
                    index=numeric.index,
                )
                return rounded.astype("Int64")
        return numeric.astype(float)

    date_signal_ratio = float(observed.map(_looks_like_date_value).mean()) if len(observed) else 0.0
    if date_signal_ratio >= 0.85 and observed.map(_looks_like_date_value).sum() >= 2:
        return pd.to_datetime(cleaned, errors="coerce", format="mixed")

    return cleaned


def _postprocess_frame(frame: pd.DataFrame) -> pd.DataFrame:
    processed = _normalise_sheet_frame(frame)
    if processed.empty:
        return processed

    processed.columns = [str(column).strip() for column in processed.columns]
    for column in processed.columns:
        processed[column] = _coerce_object_series(processed[column])
    return processed


def _is_missingish(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return bool(pd.isna(value))


def _header_text(value: Any) -> str | None:
    if _is_missingish(value):
        return None
    label = str(clean_scalar(value)).strip()
    label = re.sub(r"\s+", " ", label)
    if not label or label.lower().startswith("unnamed"):
        return None
    return label


def _row_profile(values: list[Any]) -> dict[str, float]:
    present = [value for value in values if not _is_missingish(value)]
    if not present:
        return {
            "present": 0.0,
            "textual": 0.0,
            "numeric": 0.0,
            "date": 0.0,
            "metadata": 0.0,
            "uniqueness": 0.0,
            "avg_length": 0.0,
        }

    labels = [str(clean_scalar(value)).strip() for value in present]
    numeric_like = sum(_parse_numeric_candidate(value) is not None for value in present)
    date_like = sum(_looks_like_date_value(value) for value in present)
    metadata_like = sum(
        isinstance(value, str) and METADATA_CELL_PATTERN.search(value.strip()) is not None
        for value in present
    )
    textual_like = sum(
        isinstance(value, str)
        and _parse_numeric_candidate(value) is None
        and not _looks_like_date_value(value)
        for value in present
    )
    avg_length = float(sum(len(label) for label in labels) / len(labels))
    return {
        "present": float(len(present)),
        "textual": float(textual_like),
        "numeric": float(numeric_like),
        "date": float(date_like),
        "metadata": float(metadata_like),
        "uniqueness": float(len(set(labels)) / max(len(labels), 1)),
        "avg_length": avg_length,
    }


def _is_key_value_metadata_row(values: list[Any]) -> bool:
    profile = _row_profile(values)
    if profile["present"] == 0:
        return False
    if profile["present"] <= 2 and profile["metadata"] >= 1:
        return True
    if profile["present"] <= 3 and profile["metadata"] >= 1 and profile["numeric"] + profile["date"] <= 1:
        return True
    return False


def _expand_sparse_header_tokens(tokens: list[str | None]) -> list[str | None]:
    present = sum(token is not None for token in tokens)
    if present < 2 or present >= max(2, int(len(tokens) * 0.7)):
        return tokens

    expanded: list[str | None] = []
    carry: str | None = None
    for token in tokens:
        if token is not None:
            carry = token
            expanded.append(token)
        else:
            expanded.append(carry)
    return expanded


def _combine_header_rows_multi(header_rows: list[list[Any]]) -> list[str]:
    if not header_rows:
        return []

    width = max(len(row) for row in header_rows)
    token_rows = []
    for row in header_rows:
        tokens = [_header_text(value) for value in row] + [None] * max(0, width - len(row))
        token_rows.append(_expand_sparse_header_tokens(tokens))

    combined: list[str] = []
    for index in range(width):
        pieces: list[str] = []
        for tokens in token_rows:
            token = tokens[index]
            if token and token not in pieces:
                pieces.append(token)
        combined.append(" - ".join(pieces) if pieces else f"column_{index + 1}")
    return _dedupe_columns(combined)


def _header_label_quality(labels: list[str]) -> float:
    if not labels:
        return float("-inf")

    generic = sum(label.startswith("column_") for label in labels)
    duplicates = len(labels) - len(set(labels))
    long_labels = sum(len(label) > 80 for label in labels)
    short_labels = sum(len(label.strip()) <= 1 for label in labels)
    uniqueness = len(set(labels)) / max(len(labels), 1)

    score = len(labels) * 0.8
    score += uniqueness * 5.0
    score -= generic * 1.8
    score -= duplicates * 2.0
    score -= long_labels * 1.2
    score -= short_labels * 0.8
    return score


def _data_preview_score(frame: pd.DataFrame, data_start: int) -> float:
    preview = frame.iloc[data_start : data_start + HEADER_LOOKAHEAD_ROWS]
    populated_rows = [row.tolist() for _, row in preview.iterrows() if any(not _is_missingish(v) for v in row.tolist())]
    if not populated_rows:
        return float("-inf")

    score = 0.0
    for row in populated_rows:
        profile = _row_profile(row)
        present = max(profile["present"], 1.0)
        structured_ratio = (profile["numeric"] + profile["date"]) / present
        textual_ratio = profile["textual"] / present
        score += profile["present"] * 0.35
        score += structured_ratio * 5.0
        score -= textual_ratio * 1.2
        score -= profile["metadata"] * 1.5
    return score / len(populated_rows)


def _header_candidate_score(frame: pd.DataFrame, row_index: int, depth: int) -> tuple[float, list[str], int]:
    if row_index + depth > len(frame):
        return float("-inf"), [], row_index + depth

    header_rows = [frame.iloc[idx].tolist() for idx in range(row_index, row_index + depth)]
    header_profiles = [_row_profile(row) for row in header_rows]
    if any(profile["present"] < 2 for profile in header_profiles):
        return float("-inf"), [], row_index + depth
    if _is_key_value_metadata_row(header_rows[0]):
        return float("-inf"), [], row_index + depth
    base_present = max(header_profiles[0]["present"], 1.0)
    base_text_ratio = header_profiles[0]["textual"] / base_present
    base_numeric_ratio = (header_profiles[0]["numeric"] + header_profiles[0]["date"]) / base_present
    if base_text_ratio < 0.45 or base_numeric_ratio > 0.45 or header_profiles[0]["metadata"] >= 2:
        return float("-inf"), [], row_index + depth
    if depth > 1 and any(
        not _looks_like_header_extension(frame.iloc[idx])
        for idx in range(row_index + 1, row_index + depth)
    ):
        return float("-inf"), [], row_index + depth

    labels = _combine_header_rows_multi(header_rows)
    data_start = row_index + depth
    data_score = _data_preview_score(frame, data_start)
    if data_score == float("-inf"):
        return float("-inf"), [], data_start

    header_score = sum(
        profile["present"] * 1.4
        + profile["textual"] * 1.6
        - profile["numeric"] * 1.3
        - profile["date"] * 0.9
        - profile["metadata"] * 4.5
        + profile["uniqueness"] * 2.6
        for profile in header_profiles
    )

    rows_above = frame.iloc[:row_index]
    above_bonus = 0.0
    if not rows_above.empty:
        blank_or_meta = 0
        for _, row in rows_above.iterrows():
            profile = _row_profile(row.tolist())
            if profile["present"] == 0 or profile["metadata"] >= 1:
                blank_or_meta += 1
        above_bonus = blank_or_meta * 0.4

    depth_penalty = max(0, depth - 1) * 1.2
    top_penalty = row_index * 0.15
    score = header_score + _header_label_quality(labels) + data_score + above_bonus - depth_penalty - top_penalty
    return score, labels, data_start


def _infer_sheet_layout(frame: pd.DataFrame) -> dict[str, Any]:
    normalised = _normalise_sheet_frame(frame)
    if normalised.empty:
        return {
            "header_row": 0,
            "header_depth": 1,
            "data_start_row": 1,
            "skipped_top_rows": 0,
            "confidence": 0.0,
            "columns": [],
        }

    limit = min(len(normalised), HEADER_SCAN_LIMIT)
    best: dict[str, Any] = {
        "score": float("-inf"),
        "header_row": 0,
        "header_depth": 1,
        "data_start_row": 1,
        "columns": [],
    }
    for row_index in range(limit):
        for depth in range(1, min(MAX_HEADER_DEPTH, len(normalised) - row_index) + 1):
            score, labels, data_start = _header_candidate_score(normalised, row_index, depth)
            if score > best["score"]:
                best = {
                    "score": score,
                    "header_row": row_index,
                    "header_depth": depth,
                    "data_start_row": data_start,
                    "columns": labels,
                }

    confidence = 0.0 if best["score"] == float("-inf") else max(0.0, min(1.0, (best["score"] + 8.0) / 55.0))
    return {
        "header_row": int(best["header_row"]),
        "header_depth": int(best["header_depth"]),
        "data_start_row": int(best["data_start_row"]),
        "skipped_top_rows": int(best["header_row"]),
        "confidence": round(confidence, 4),
        "columns": best["columns"],
    }


def _header_score(frame: pd.DataFrame, row_index: int) -> float:
    row = frame.iloc[row_index].tolist()
    present = [value for value in row if not _is_missingish(value)]
    if len(present) < 2:
        return float("-inf")

    as_text = [str(value).strip() for value in present]
    string_like = sum(isinstance(value, str) for value in present)
    numeric_like = sum(
        isinstance(value, (int, float, np.integer, np.floating, pd.Timestamp, datetime))
        for value in present
    )
    metadata_like = sum(
        isinstance(value, str) and METADATA_CELL_PATTERN.search(value.strip()) is not None
        for value in present
    )
    uniqueness = len(set(as_text)) / max(len(as_text), 1)

    next_rows = frame.iloc[row_index + 1 : row_index + 1 + HEADER_LOOKAHEAD_ROWS]
    next_non_empty = 0.0
    object_heavy = 0
    for _, next_row in next_rows.iterrows():
        values = [value for value in next_row.tolist() if not _is_missingish(value)]
        if not values:
            continue
        next_non_empty += len(values)
        object_heavy += int(sum(isinstance(value, str) for value in values) >= max(1, len(values) / 2))

    score = 0.0
    score += len(present) * 2.0
    score += string_like * 1.4
    score -= numeric_like * 0.8
    score += uniqueness * 3.0
    score += next_non_empty * 0.12
    score += object_heavy * 0.4
    score -= metadata_like * 4.0
    if len(present) == 1:
        score -= 6.0
    if all(isinstance(value, str) and len(str(value).strip()) <= 2 for value in present):
        score -= 2.0
    return score


def _best_header_row(frame: pd.DataFrame) -> int:
    if frame.empty:
        return 0
    limit = min(len(frame), HEADER_SCAN_LIMIT)
    best_index = 0
    best_score = float("-inf")
    for row_index in range(limit):
        score = _header_score(frame, row_index)
        if score > best_score:
            best_index = row_index
            best_score = score
    return best_index


def _clean_header_value(value: Any, index: int) -> str:
    if _is_missingish(value):
        return f"column_{index + 1}"
    label = str(clean_scalar(value)).strip()
    label = re.sub(r"\s+", " ", label)
    if not label or label.lower().startswith("unnamed"):
        return f"column_{index + 1}"
    return label


def _dedupe_columns(values: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    columns: list[str] = []
    for value in values:
        count = counts.get(value, 0) + 1
        counts[value] = count
        columns.append(value if count == 1 else f"{value}_{count}")
    return columns


def _looks_like_header_extension(row: pd.Series) -> bool:
    values = [value for value in row.tolist() if not _is_missingish(value)]
    if len(values) < 2:
        return False
    if _is_key_value_metadata_row(values):
        return False
    string_like = sum(isinstance(value, str) for value in values)
    numeric_like = sum(_parse_numeric_candidate(value) is not None for value in values)
    date_like = sum(_looks_like_date_value(value) for value in values)
    textual_like = len(values) - numeric_like - date_like
    return (
        string_like >= max(2, len(values) * 0.6)
        and textual_like >= max(2, len(values) * 0.5)
        and numeric_like <= max(1, int(len(values) * 0.2))
    )


def _combine_header_rows(primary: list[Any], secondary: list[Any]) -> list[Any]:
    combined: list[Any] = []
    for index, value in enumerate(primary):
        upper = _clean_header_value(value, index)
        lower_raw = secondary[index] if index < len(secondary) else None
        lower = None if _is_missingish(lower_raw) else str(clean_scalar(lower_raw)).strip()
        if lower and lower.lower() != "unnamed" and lower != upper and not upper.startswith("column_"):
            combined.append(f"{upper}-{lower}")
        elif lower and upper.startswith("column_"):
            combined.append(lower)
        else:
            combined.append(upper)
    return combined


def _prepare_excel_sheet(frame: pd.DataFrame) -> pd.DataFrame:
    normalised = _normalise_sheet_frame(frame)
    if normalised.empty:
        return normalised

    layout = _infer_sheet_layout(normalised)
    columns = layout["columns"] or [
        _clean_header_value(value, index)
        for index, value in enumerate(normalised.iloc[layout["header_row"]].tolist())
    ]
    prepared = normalised.iloc[layout["data_start_row"] :].copy()
    prepared.columns = columns
    prepared = prepared.dropna(how="all").reset_index(drop=True)
    return _postprocess_frame(prepared)


def list_datasets() -> list[dict[str, Any]]:
    ensure_storage()
    now_ts = datetime.now(timezone.utc).timestamp()
    with _DATASET_LIST_CACHE_LOCK:
        if (
            now_ts < float(_DATASET_LIST_CACHE.get("expires_at") or 0.0)
            and isinstance(_DATASET_LIST_CACHE.get("items"), list)
        ):
            return [dict(item) for item in list(_DATASET_LIST_CACHE.get("items") or [])]

    metadata_paths = sorted(DATASETS_DIR.glob("*/metadata.json"))
    signature = tuple(
        (
            str(path),
            int(path.stat().st_mtime_ns) if path.exists() else 0,
            int(path.stat().st_size) if path.exists() else 0,
        )
        for path in metadata_paths
    )
    with _DATASET_LIST_CACHE_LOCK:
        if (
            _DATASET_LIST_CACHE.get("signature") == signature
            and now_ts < float(_DATASET_LIST_CACHE.get("expires_at") or 0.0)
            and isinstance(_DATASET_LIST_CACHE.get("items"), list)
        ):
            return [dict(item) for item in list(_DATASET_LIST_CACHE.get("items") or [])]

    items: list[dict[str, Any]] = []
    for metadata_path in metadata_paths:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if isinstance(metadata, dict):
            items.append(_dataset_list_item(metadata))
    items.sort(key=lambda item: item.get("last_updated", ""), reverse=True)
    with _DATASET_LIST_CACHE_LOCK:
        _DATASET_LIST_CACHE.update(
            {
                "expires_at": now_ts + DATASET_LIST_CACHE_TTL_SECONDS,
                "signature": signature,
                "items": [dict(item) for item in items],
            }
        )
    return items


def load_dataset_metadata(dataset_id: str) -> dict[str, Any]:
    metadata_path = DATASETS_DIR / dataset_id / "metadata.json"
    if not metadata_path.exists():
        raise HTTPException(status_code=404, detail="Dataset not found.")
    return json.loads(metadata_path.read_text(encoding="utf-8"))


def load_dataset_frame(dataset_id: str, sheet_name: str | None = None) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]]:
    metadata = load_dataset_metadata(dataset_id)
    active_sheet_name = sheet_name or metadata.get("active_sheet")
    sheets = metadata.get("sheets", [])
    sheet = next((item for item in sheets if item["name"] == active_sheet_name), None)
    if sheet is None and sheets:
        sheet = sheets[0]
    if sheet is None:
        raise HTTPException(status_code=400, detail="Dataset has no readable sheet.")

    dataset_dir = DATASETS_DIR / dataset_id
    frame_path = dataset_dir / sheet["storage_file"]
    if not frame_path.exists():
        raise HTTPException(status_code=404, detail="Stored dataset file is missing.")

    frame = _postprocess_frame(pd.read_pickle(frame_path))
    return frame, metadata, sheet


def load_all_sheet_frames(dataset_id: str) -> tuple[dict[str, Any], list[tuple[dict[str, Any], pd.DataFrame]]]:
    metadata = load_dataset_metadata(dataset_id)
    dataset_dir = DATASETS_DIR / dataset_id
    sheets: list[tuple[dict[str, Any], pd.DataFrame]] = []
    for sheet in metadata.get("sheets", []):
        frame_path = dataset_dir / sheet["storage_file"]
        if frame_path.exists():
            sheets.append((sheet, _postprocess_frame(pd.read_pickle(frame_path))))
    return metadata, sheets


def numeric_columns(frame: pd.DataFrame) -> list[str]:
    return frame.select_dtypes(include=["number", "bool"]).columns.astype(str).tolist()


def datetime_columns(frame: pd.DataFrame) -> list[str]:
    return [
        str(column)
        for column in frame.columns
        if pd.api.types.is_datetime64_any_dtype(frame[column])
    ]


def categorical_columns(frame: pd.DataFrame, numeric: list[str], datetimes: list[str]) -> list[str]:
    items: list[str] = []
    for column in frame.columns.astype(str).tolist():
        if column in numeric or column in datetimes:
            continue
        if frame[column].nunique(dropna=True) <= 80:
            items.append(column)
    return items


def build_column_summaries(frame: pd.DataFrame) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    row_count = max(len(frame), 1)
    for column in frame.columns.astype(str).tolist():
        series = frame[column]
        summary: dict[str, Any] = {
            "name": column,
            "dtype": str(series.dtype),
            "missing_count": int(series.isna().sum()),
            "missing_ratio": float(series.isna().sum() / row_count),
            "unique_count": int(series.nunique(dropna=True)),
            "sample_values": [clean_scalar(value) for value in series.dropna().head(4).tolist()],
        }
        if pd.api.types.is_numeric_dtype(series):
            summary["stats"] = {
                "mean": clean_scalar(series.mean()),
                "std": clean_scalar(series.std()),
                "min": clean_scalar(series.min()),
                "max": clean_scalar(series.max()),
            }
        else:
            top_values = series.fillna("Missing").astype(str).value_counts().head(3)
            summary["stats"] = {
                "top_values": [
                    {"label": label, "count": int(count)}
                    for label, count in top_values.items()
                ]
            }
        summaries.append(summary)
    return summaries


def build_chart_bundle(frame: pd.DataFrame) -> dict[str, Any]:
    numeric = numeric_columns(frame)
    datetimes = datetime_columns(frame)
    categorical = categorical_columns(frame, numeric, datetimes)
    charts: dict[str, Any] = {}

    if numeric:
        primary = numeric[0]
        values = pd.to_numeric(frame[primary], errors="coerce").dropna()
        if not values.empty:
            bins = min(12, max(5, int(np.sqrt(len(values)))))
            counts, edges = np.histogram(values, bins=bins)
            charts["distribution"] = {
                "kind": "histogram",
                "title": f"{primary} distribution",
                "x": [f"{edges[idx]:.2f} - {edges[idx + 1]:.2f}" for idx in range(len(edges) - 1)],
                "y": counts.tolist(),
            }

    if categorical:
        primary = categorical[0]
        counts = frame[primary].fillna("Missing").astype(str).value_counts().head(10)
        charts["category"] = {
            "kind": "bar",
            "title": f"{primary} mix",
            "x": counts.index.tolist(),
            "y": [int(value) for value in counts.tolist()],
        }

    if len(numeric) >= 2:
        corr_frame = frame[numeric[:8]].corr(numeric_only=True).fillna(0)
        charts["correlation"] = {
            "kind": "heatmap",
            "title": "Correlation map",
            "labels": corr_frame.columns.astype(str).tolist(),
            "matrix": corr_frame.round(4).to_numpy().tolist(),
        }

        scatter = frame[numeric[:2]].dropna().head(240)
        charts["scatter"] = {
            "kind": "scatter",
            "title": f"{numeric[0]} vs {numeric[1]}",
            "x_label": numeric[0],
            "y_label": numeric[1],
            "points": [
                [clean_scalar(x), clean_scalar(y)]
                for x, y in scatter.to_numpy().tolist()
            ],
        }

    return charts


def build_profile(
    *,
    dataset_id: str,
    name: str,
    filename: str,
    active_sheet: str,
    sheets: list[dict[str, Any]],
    frame: pd.DataFrame,
) -> dict[str, Any]:
    profile_frame, profile_sample_strategy = _profile_sample_frame(frame)
    profile_mode = "full" if profile_sample_strategy == "full_dataset" else "sampled"
    numeric = numeric_columns(frame)
    datetimes = datetime_columns(frame)
    categorical = categorical_columns(frame, numeric, datetimes)
    profile_numeric = numeric_columns(profile_frame)
    profile_datetimes = datetime_columns(profile_frame)
    profile_categorical = categorical_columns(profile_frame, profile_numeric, profile_datetimes)
    return {
        "dataset_id": dataset_id,
        "name": name,
        "filename": filename,
        "active_sheet": active_sheet,
        "sheets": sheets,
        "row_count": int(len(frame)),
        "column_count": int(len(frame.columns)),
        "numeric_columns": numeric,
        "categorical_columns": categorical,
        "datetime_columns": datetimes,
        "missing_cells": int(frame.isna().sum().sum()),
        "sample_rows": clean_records(profile_frame),
        "column_summaries": build_column_summaries(profile_frame),
        "chart_bundle": build_chart_bundle(profile_frame),
        "enhanced_data_profile": {
            "contract": "asteria_enhanced_data_profile_v1",
            "profile_mode": profile_mode,
            "full_row_count": int(len(frame)),
            "profile_row_count": int(len(profile_frame)),
            "profile_sample_strategy": profile_sample_strategy,
            "profile_sample_row_limit": PROFILE_SAMPLE_ROW_LIMIT,
            "large_dataset_threshold": LARGE_DATASET_ROW_THRESHOLD,
            "large_dataset": bool(len(frame) >= LARGE_DATASET_ROW_THRESHOLD),
            "full_numeric_column_count": len(numeric),
            "full_categorical_column_count": len(categorical),
            "profile_numeric_column_count": len(profile_numeric),
            "profile_categorical_column_count": len(profile_categorical),
            "capability": "Full rows are persisted for analysis; browser-facing profile/chart previews are sampled to keep large imports responsive.",
        },
        "last_updated": utc_now(),
    }


async def persist_dataset(upload_file: UploadFile) -> dict[str, Any]:
    ensure_storage()
    filename = Path(upload_file.filename or "dataset").name
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx, .csv, .tsv, and .dta files are supported in this build.",
        )

    dataset_id = uuid.uuid4().hex[:12]
    dataset_dir = DATASETS_DIR / dataset_id
    dataset_dir.mkdir(parents=True, exist_ok=True)

    raw_path = dataset_dir / f"source{suffix}"
    raw_path.write_bytes(await upload_file.read())

    dataset_name = Path(filename).stem.replace("_", " ").strip() or f"Dataset {dataset_id}"
    sheets: list[dict[str, Any]] = []

    if suffix == ".xlsx":
        workbook_sheets = _load_excel_workbook_sheets(raw_path)
        for index, (sheet_name, raw_frame, merge_metadata) in enumerate(workbook_sheets):
            layout = _infer_sheet_layout(raw_frame)
            frame = _prepare_excel_sheet(raw_frame)
            storage_file = f"{slugify(sheet_name or f'sheet-{index + 1}')}.pkl"
            frame.to_pickle(dataset_dir / storage_file)
            sheets.append(
                {
                    "name": sheet_name,
                    "storage_file": storage_file,
                    "rows": int(len(frame)),
                    "columns": int(len(frame.columns)),
                    "layout": layout,
                    **merge_metadata,
                }
            )
    elif suffix == ".dta":
        frame = _postprocess_frame(pd.read_stata(raw_path, convert_categoricals=False))
        storage_file = "sheet-main.pkl"
        frame.to_pickle(dataset_dir / storage_file)
        sheets.append(
            {
                "name": "Sheet1",
                "storage_file": storage_file,
                "rows": int(len(frame)),
                "columns": int(len(frame.columns)),
            }
        )
    else:
        sep = "\t" if suffix == ".tsv" else ","
        frame = _postprocess_frame(pd.read_csv(raw_path, sep=sep))
        storage_file = "sheet-main.pkl"
        frame.to_pickle(dataset_dir / storage_file)
        sheets.append(
            {
                "name": "Sheet1",
                "storage_file": storage_file,
                "rows": int(len(frame)),
                "columns": int(len(frame.columns)),
            }
        )

    if not sheets:
        raise HTTPException(status_code=400, detail="No readable sheets were found in the uploaded file.")

    active_sheet = sheets[0]["name"]
    active_frame = pd.read_pickle(dataset_dir / sheets[0]["storage_file"])
    metadata = build_profile(
        dataset_id=dataset_id,
        name=dataset_name,
        filename=filename,
        active_sheet=active_sheet,
        sheets=sheets,
        frame=active_frame,
    )
    save_metadata(dataset_dir, metadata)
    return metadata


def activate_sheet(dataset_id: str, sheet_name: str) -> dict[str, Any]:
    metadata = load_dataset_metadata(dataset_id)
    available = {sheet["name"] for sheet in metadata.get("sheets", [])}
    if sheet_name not in available:
        raise HTTPException(status_code=404, detail="Sheet not found for this dataset.")

    frame, _, _ = load_dataset_frame(dataset_id, sheet_name)
    updated = build_profile(
        dataset_id=metadata["dataset_id"],
        name=metadata["name"],
        filename=metadata["filename"],
        active_sheet=sheet_name,
        sheets=metadata["sheets"],
        frame=frame,
    )
    save_metadata(DATASETS_DIR / dataset_id, updated)
    return updated
