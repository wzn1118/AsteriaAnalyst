from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd

from app.models import SmartReportRequest
from app.services.r_workflow_service import (
    _build_r_workflow_column_registry,
    _build_interpretation_markdown,
    _build_method_analysis_cards,
    _run_r_workflow_codex_sidecar,
    _write_r_workflow_statistics_workbook,
    run_r_workflow,
)


def _fake_rscript(
    runtime_path: str,
    script_path: Path,
    args: list[str],
    workflow_dir: Path,
    *,
    log_prefix: str,
) -> None:
    """Provide deterministic R-boundary artifacts for unit-level workflow tests.

    R is an optional user-configured runtime. These tests exercise Asteria's
    artifact handling and report contract, while real R execution remains an
    opt-in integration concern.
    """
    del runtime_path, script_path
    (workflow_dir / f"{log_prefix}-stdout.log").write_text("fake R runtime\n", encoding="utf-8")
    (workflow_dir / f"{log_prefix}-stderr.log").write_text("", encoding="utf-8")

    if log_prefix == "01_clean_prepare":
        pd.read_csv(args[0], encoding="utf-8-sig").to_csv(
            workflow_dir / "cleaned_dataset.csv", index=False, encoding="utf-8-sig"
        )
        return

    frame = pd.read_csv(workflow_dir / "input-data.csv", encoding="utf-8-sig")
    numeric_columns = [str(column) for column in frame.select_dtypes(include="number").columns]
    variable_columns = [
        column for column in numeric_columns if frame[column].dropna().nunique() > 1
    ]

    summary_rows = [
        {
            "column": column,
            "n": int(frame[column].notna().sum()),
            "mean": float(frame[column].mean()),
        }
        for column in numeric_columns
    ]
    pd.DataFrame(summary_rows, columns=["column", "n", "mean"]).to_csv(
        workflow_dir / "summary_stats.csv", index=False, encoding="utf-8-sig"
    )

    category_columns = [
        str(column)
        for column in frame.columns
        if str(frame[column].dtype) in {"object", "string"} and str(column).lower() != "date"
    ]
    if category_columns:
        category = category_columns[0]
        top_categories = (
            frame[category]
            .fillna("(missing)")
            .astype(str)
            .value_counts()
            .rename_axis("category")
            .reset_index(name="count")
        )
    else:
        top_categories = pd.DataFrame(columns=["category", "count"])
    top_categories.to_csv(workflow_dir / "top_categories.csv", index=False, encoding="utf-8-sig")

    pd.DataFrame(columns=["period", "metric", "value"]).to_csv(
        workflow_dir / "temporal_trend.csv", index=False, encoding="utf-8-sig"
    )
    pd.DataFrame(columns=["period_from", "period_to", "metric", "growth_rate", "delta"]).to_csv(
        workflow_dir / "temporal_growth.csv", index=False, encoding="utf-8-sig"
    )
    pd.DataFrame(columns=["variable", "correlation"]).to_csv(
        workflow_dir / "correlation_matrix.csv", index=False, encoding="utf-8-sig"
    )

    pd.DataFrame(
        [
            {"method": "summary_stats", "output": "summary_stats.csv", "status": True, "note": "fixture"},
            {"method": "temporal_growth", "output": "temporal_growth.csv", "status": True, "note": "single-period-safe"},
            {"method": "pca_variance", "output": "pca_axis_summary.csv", "status": True, "note": "variable-columns-only"},
            {"method": "kmeans_clusters", "output": "cluster_member_detail.csv", "status": True, "note": "fixture"},
        ]
    ).to_csv(workflow_dir / "method_log.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [{"component": "PC1", "explained_variance": 1.0}]
        if variable_columns
        else [],
        columns=["component", "explained_variance"],
    ).to_csv(workflow_dir / "pca_axis_summary.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [
            {"variable": column, "PC1": 1.0 / (index + 1), "PC2": 0.0}
            for index, column in enumerate(variable_columns)
        ],
        columns=["variable", "PC1", "PC2"],
    ).to_csv(workflow_dir / "pca_loadings.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [{"row_index": index + 1, "cluster": (index % 2) + 1} for index in range(len(frame))]
    ).to_csv(workflow_dir / "cluster_member_detail.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(
        [
            {"category": "all", "metric": column, "mean": float(frame[column].mean())}
            for column in numeric_columns
        ],
        columns=["category", "metric", "mean"],
    ).to_csv(workflow_dir / "category_metric_summary.csv", index=False, encoding="utf-8-sig")
    if len(numeric_columns) >= 2:
        gap = float(frame[numeric_columns[0]].mean() - frame[numeric_columns[1]].mean())
        pair = f"{numeric_columns[0]} / {numeric_columns[1]}"
        budget_rows = [{"metric_pair": pair, "gap": gap}]
    else:
        budget_rows = []
    pd.DataFrame(budget_rows, columns=["metric_pair", "gap"]).to_csv(
        workflow_dir / "budget_variance_summary.csv", index=False, encoding="utf-8-sig"
    )


class RWorkflowServiceTests(unittest.TestCase):
    def test_r_workflow_codex_sidecar_prefers_runtime_task_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            with patch("app.services.r_workflow_service.load_runtime_settings_raw", return_value={"codex_runtime_enabled": True, "codex_search_enabled": False, "codex_timeout_sec": 900, "codex_dangerously_bypass_approvals_and_sandbox": False}), patch(
                "app.services.r_workflow_service.create_codex_run_task",
                return_value={
                    "job_id": "codex-task-sidecar-001",
                    "run_id": "",
                    "status": "queued",
                    "parent_report_job_id": "report-task-001",
                    "parent_report_id": "report-001",
                    "stage_id": "r_workflow_sidecar",
                    "purpose": "r_workflow_sidecar_review",
                },
            ), patch(
                "app.services.r_workflow_service.get_codex_run_task",
                return_value={
                    "job_id": "codex-task-sidecar-001",
                    "run_id": "run-sidecar-001",
                    "status": "running",
                    "result_summary": {"summary": "runtime task accepted"},
                    "error": "",
                },
            ), patch("app.services.r_workflow_service.run_headless_codex") as mocked_sync:
                result = _run_r_workflow_codex_sidecar(
                    workflow_dir=workflow_dir,
                    report_id="report-001",
                    report_job_id="report-task-001",
                    dataset_name="demo-dataset",
                    sheet_name="Sheet1",
                    report_lens="generic_business_review",
                    request=SmartReportRequest(sheet_name="Sheet1", user_requirement="review R workflow"),
                    interpretation_summary=["summary"],
                    column_role_registry={"numeric_method_columns": ["gmv"]},
                )

        self.assertEqual(result["task_id"], "codex-task-sidecar-001")
        self.assertEqual(result["run_id"], "run-sidecar-001")
        self.assertEqual(result["status"], "running")
        self.assertEqual(result["reason"], "")
        mocked_sync.assert_not_called()

    def test_r_workflow_codex_sidecar_falls_back_to_sync_when_task_mode_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            with patch("app.services.r_workflow_service.load_runtime_settings_raw", return_value={"codex_runtime_enabled": True, "codex_search_enabled": False, "codex_timeout_sec": 900, "codex_dangerously_bypass_approvals_and_sandbox": False}), patch(
                "app.services.r_workflow_service.create_codex_run_task",
                side_effect=RuntimeError("task create failed"),
            ), patch(
                "app.services.r_workflow_service.run_headless_codex",
                return_value={
                    "status": "completed",
                    "run_id": "run-sync-001",
                    "session_id": "session-sync-001",
                    "summary": "sync fallback ok",
                    "changed_files": [],
                    "transcript_entry_count": 3,
                    "error": "",
                    "transcript_path": "",
                    "transcript_url": "",
                    "stdout_path": "",
                    "stdout_url": "",
                    "stderr_path": "",
                    "stderr_url": "",
                    "summary_path": "",
                    "summary_url": "",
                    "git_diff_path": "",
                    "git_diff_url": "",
                },
            ):
                result = _run_r_workflow_codex_sidecar(
                    workflow_dir=workflow_dir,
                    report_id="report-002",
                    report_job_id="report-task-002",
                    dataset_name="demo-dataset",
                    sheet_name="Sheet1",
                    report_lens="generic_business_review",
                    request=SmartReportRequest(sheet_name="Sheet1", user_requirement="review R workflow"),
                    interpretation_summary=["summary"],
                    column_role_registry={"numeric_method_columns": ["gmv"]},
                )

        self.assertEqual(result["task_id"], "")
        self.assertEqual(result["run_id"], "run-sync-001")
        self.assertTrue(result["reason"].startswith("task_mode_failed:"))

    def test_column_registry_keeps_numeric_metrics_and_excludes_name_date(self) -> None:
        import pandas as pd

        frame = pd.DataFrame(
            {
                "name": ["a", "b", "c", "d"],
                "date": pd.to_datetime(["2026-04-01", "2026-04-02", "2026-04-03", "2026-04-04"]),
                "gmv": [1000, 1200, 900, 1400],
                "orders": [10, 12, 9, 14],
                "users": [100, 115, 98, 130],
                "category": ["A", "B", "A", "B"],
            }
        )

        registry = _build_r_workflow_column_registry(frame)
        self.assertIn("gmv", registry["numeric_method_columns"])
        self.assertIn("orders", registry["numeric_method_columns"])
        self.assertIn("users", registry["numeric_method_columns"])
        self.assertNotIn("name", registry["numeric_method_columns"])
        self.assertNotIn("date", registry["numeric_method_columns"])
        self.assertIn("date", registry["temporal_columns"])
        self.assertIn("category", registry["category_dimension_columns"])

    def test_column_registry_infers_variance_pairs_without_exact_hardcoded_names(self) -> None:
        import pandas as pd

        frame = pd.DataFrame(
            {
                "月份": ["2026-01", "2026-02", "2026-03", "2026-04"],
                "部门": ["A", "A", "B", "B"],
                "计划收入": [100, 110, 120, 130],
                "实际收入额": [102, 115, 118, 135],
                "预算费用": [40, 42, 45, 48],
                "实际费用": [41, 43, 47, 49],
            }
        )

        registry = _build_r_workflow_column_registry(frame)
        pair_markers = {f"{item['baseline_column']}::{item['actual_column']}" for item in registry["variance_pairs"]}
        self.assertIn("计划收入::实际收入额", pair_markers)
        self.assertIn("预算费用::实际费用", pair_markers)

    def test_method_cards_cover_all_summary_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "summary_stats.csv").write_text(
                "\n".join(
                    [
                        '"column","n","mean","median","sd","min","max"',
                        '"gmv",4,1125,1100,221.7,900,1400',
                        '"orders",4,11.25,11,2.2,9,14',
                        '"users",4,110.75,107.5,14.9,98,130',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "summary_stats",
                        "output": "summary_stats.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            self.assertEqual(len(cards), 1)
            self.assertEqual(len(cards[0]["current_readings"]), 3)
            joined = " ".join(cards[0]["current_readings"])
            self.assertIn("`gmv`", joined)
            self.assertIn("`orders`", joined)
            self.assertIn("`users`", joined)

    def test_method_cards_cover_multiple_category_metric_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "category_metric_summary.csv").write_text(
                "\n".join(
                    [
                        '"dimension","category","metric","mean_value"',
                        '"category","A","gmv",950',
                        '"category","B","gmv",1300',
                        '"category","A","orders",9.5',
                        '"category","B","orders",13',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "category_metric_summary",
                        "output": "category_metric_summary.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            self.assertEqual(len(cards), 1)
            joined = " ".join(cards[0]["current_readings"])
            self.assertIn("`gmv`", joined)
            self.assertIn("`orders`", joined)

    def test_temporal_method_cards_cover_all_metrics(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "temporal_trend.csv").write_text(
                "\n".join(
                    [
                        '"period","metric","value"',
                        '2026-04-01,"gmv",1000',
                        '2026-04-02,"gmv",1200',
                        '2026-04-01,"orders",10',
                        '2026-04-02,"orders",12',
                        '2026-04-01,"users",100',
                        '2026-04-02,"users",115',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "temporal_trend",
                        "output": "temporal_trend.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            self.assertEqual(len(cards), 1)
            self.assertEqual(len(cards[0]["current_readings"]), 3)
            joined = " ".join(cards[0]["current_readings"])
            self.assertIn("`gmv`", joined)
            self.assertIn("`orders`", joined)
            self.assertIn("`users`", joined)

    def test_pca_variance_method_card_uses_actual_csv_column_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "pca_variance.csv").write_text(
                "\n".join(
                    [
                        '"component","std_dev","variance_explained","cumulative_variance"',
                        '"PC1",1.9,0.711,0.711',
                        '"PC2",0.8,0.135,0.846',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "pca_variance",
                        "output": "pca_variance.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertNotIn("n/a", joined.lower())
            self.assertIn("`PC1`", joined)
            self.assertIn("解释方差 0.7", joined)

    def test_pca_loadings_method_card_reads_pc_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "pca_loadings.csv").write_text(
                "\n".join(
                    [
                        '"variable","PC1","PC2"',
                        '"page_views",0.71,0.12',
                        '"cart_count",0.69,0.15',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "pca_loadings",
                        "output": "pca_loadings.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertNotIn("n/a", joined.lower())
            self.assertIn("`PC1` - `page_views`", joined)
            self.assertIn("`PC2` - `cart_count`", joined)

    def test_kmeans_clusters_method_card_counts_cluster_sizes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "kmeans_clusters.csv").write_text(
                "\n".join(
                    [
                        '"index","cluster"',
                        '1,2',
                        '2,2',
                        '3,1',
                        '4,3',
                        '5,2',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "kmeans_clusters",
                        "output": "kmeans_clusters.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertNotIn("n/a", joined.lower())
            self.assertIn("`2`：样本量 3", joined)
            self.assertIn("`1`：样本量 1", joined)

    def test_cluster_profile_method_card_compares_clusters_to_overall(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "summary_stats.csv").write_text(
                "\n".join(
                    [
                        '"column","n","mean","median","sd","min","max"',
                        '"Revenue",5,100,90,10,80,130',
                        '"GMV",5,120,110,12,90,150',
                        '"FreightCost",5,20,18,4,10,30',
                    ]
                ),
                encoding="utf-8",
            )
            (workflow_dir / "kmeans_clusters.csv").write_text(
                "\n".join(
                    [
                        '"index","cluster"',
                        '1,1',
                        '2,1',
                        '3,1',
                        '4,2',
                        '5,2',
                    ]
                ),
                encoding="utf-8",
            )
            (workflow_dir / "cluster_profile.csv").write_text(
                "\n".join(
                    [
                        '"cluster","Revenue","GMV","FreightCost"',
                        '1,180,210,28',
                        '2,60,70,12',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "cluster_profile",
                        "output": "cluster_profile.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertIn("`1`", joined)
            self.assertIn("`Revenue`偏高", joined)
            self.assertIn("主流群体", joined)

    def test_pca_scores_method_card_summarizes_score_ranges(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "pca_scores.csv").write_text(
                "\n".join(
                    [
                        '"index","PC1","PC2"',
                        '1,-2.0,-1.0',
                        '2,-1.0,-0.5',
                        '3,0.0,0.0',
                        '4,1.0,0.5',
                        '5,3.0,1.5',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "pca_scores",
                        "output": "pca_scores.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertIn("`PC1` 得分主要落在", joined)
            self.assertIn("远离主体", joined)

    def test_top_categories_method_card_does_not_require_share(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "top_categories.csv").write_text(
                "\n".join(
                    [
                        '"dimension","category","count"',
                        '"category_id","4756105",9124',
                        '"category_id","4145813",9001',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "top_categories",
                        "output": "top_categories.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertNotIn("n/a", joined.lower())
            self.assertIn("数量 9,124", joined)

    def test_category_price_band_method_card_reads_band_and_count(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "category_price_band_summary.csv").write_text(
                "\n".join(
                    [
                        '"dimension","category","metric","band","count"',
                        '"category_id","2171","page_views","[1,2]",8',
                        '"category_id","2410","page_views","[1,2]",5',
                    ]
                ),
                encoding="utf-8",
            )

            cards = _build_method_analysis_cards(
                workflow_dir,
                [
                    {
                        "method": "category_price_band_summary",
                        "output": "category_price_band_summary.csv",
                        "status": "TRUE",
                        "note": "",
                    }
                ],
            )

            joined = " ".join(cards[0]["current_readings"])
            self.assertNotIn("n/a", joined.lower())
            self.assertIn("分位带 `[1,2]`", joined)
            self.assertIn("样本量 8", joined)

    def test_markdown_renders_multiple_method_readings(self) -> None:
        method_cards = [
            {
                "title": "数值摘要",
                "status": "已执行",
                "what": "看整体数值特征",
                "current_reading": "fallback",
                "current_readings": [
                    "`gmv`：均值 1125。",
                    "`orders`：均值 11.25。",
                    "`users`：均值 110.75。",
                ],
                "business_use": "用于判断盘面规模",
                "next_action": "继续结合趋势看变化",
            }
        ]

        markdown = _build_interpretation_markdown(
            dataset_name="demo",
            sheet_name="Sheet1",
            request=SmartReportRequest(
                sheet_name="Sheet1",
                use_r_workflow=True,
                core_purpose="验证方法逐项解读",
            ),
            runtime_path="Rscript.exe",
            interpretation_summary=["summary"],
            interpretation_sections=[],
            method_cards=method_cards,
            output_rows=[],
            csv_sections=[],
            code_sections=[],
            image_sections=[],
            log_sections=[],
        )

        self.assertIn("- 当前看到什么：", markdown)
        self.assertIn("  - `gmv`：均值 1125。", markdown)
        self.assertIn("  - `orders`：均值 11.25。", markdown)
        self.assertIn("  - `users`：均值 110.75。", markdown)

    def test_statistics_workbook_collects_r_csv_outputs(self) -> None:
        try:
            from openpyxl import load_workbook
        except Exception:
            self.skipTest("openpyxl unavailable")

        import pandas as pd

        with tempfile.TemporaryDirectory() as tmp:
            workflow_dir = Path(tmp)
            (workflow_dir / "summary_stats.csv").write_text(
                "\n".join(
                    [
                        '"column","n","mean"',
                        '"gmv",4,1125',
                        '"orders",4,11.25',
                    ]
                ),
                encoding="utf-8",
            )
            (workflow_dir / "method_log.csv").write_text(
                "\n".join(
                    [
                        '"method","output","status","note"',
                        '"summary_stats","summary_stats.csv",TRUE,"ok"',
                    ]
                ),
                encoding="utf-8",
            )
            registry = _build_r_workflow_column_registry(
                pd.DataFrame(
                    {
                        "date": pd.to_datetime(["2026-04-01", "2026-04-02", "2026-04-03", "2026-04-04"]),
                        "category": ["A", "B", "A", "B"],
                        "gmv": [1000, 1200, 900, 1400],
                        "orders": [10, 12, 9, 14],
                    }
                )
            )
            workbook_path = _write_r_workflow_statistics_workbook(
                workflow_dir=workflow_dir,
                report_id="demoexcel01",
                dataset_name="demo",
                sheet_name="Sheet1",
                runtime_path="Rscript.exe",
                column_role_registry=registry,
            )

            workbook = load_workbook(workbook_path)
            self.assertIn("overview", workbook.sheetnames)
            self.assertIn("results_index", workbook.sheetnames)
            self.assertIn("column_roles", workbook.sheetnames)
            self.assertIn("summary_stats", workbook.sheetnames)
            self.assertIn("method_log", workbook.sheetnames)

    def test_single_day_large_sample_skips_temporal_growth_without_crashing(self) -> None:
        import shutil
        import uuid
        from unittest.mock import patch

        import app.services.r_workflow_service as rws
        import app.services.codex_service as cs
        from app.services.path_service import REPORTS_DIR
        import pandas as pd

        frame = pd.DataFrame(
            {
                "date": ["2017-12-02"] * 12,
                "item_id": [f"i{i}" for i in range(12)],
                "category_id": ["c1", "c2"] * 6,
                "page_views": [100, 120, 80, 95, 140, 160, 110, 130, 90, 105, 150, 170],
                "cart_count": [10, 12, 8, 9, 14, 16, 11, 13, 9, 10, 15, 17],
                "fav_count": [5, 6, 4, 4, 7, 8, 5, 6, 4, 5, 7, 8],
                "buy_count": [2, 3, 1, 2, 4, 5, 3, 3, 2, 2, 4, 5],
                "user_count": [70, 82, 65, 68, 90, 96, 75, 80, 69, 71, 92, 99],
                "buy_rate": [0.02, 0.025, 0.0125, 0.021, 0.028, 0.031, 0.027, 0.023, 0.022, 0.019, 0.026, 0.029],
                "cart_rate": [0.10, 0.10, 0.10, 0.095, 0.10, 0.10, 0.10, 0.10, 0.10, 0.095, 0.10, 0.10],
            }
        )

        report_id = "singleday" + uuid.uuid4().hex[:8]
        report_dir = REPORTS_DIR / f"smart-report-{report_id}"
        if report_dir.exists():
            shutil.rmtree(report_dir)

        with patch.object(rws, "codex_generate_r_workflow", lambda ctx: cs._fallback_r_workflow_author_v2("forced-test", ctx)), patch.object(
            rws, "codex_interpret_r_results", lambda payload: cs._fallback_r_results_interpretation("forced-test", payload)
        ), patch.object(
            rws, "codex_infer_r_workflow_semantics", lambda ctx: ctx["heuristic_registry"]
        ), patch.object(
            rws, "_resolve_rscript_path", return_value="fake-rscript"
        ), patch.object(
            rws, "_run_rscript", side_effect=_fake_rscript
        ), patch.object(
            rws, "_run_r_workflow_codex_sidecar", return_value={"enabled": False, "downloadables": []}
        ):
            result = run_r_workflow(
                report_dir=report_dir,
                report_id=report_id,
                dataset_name="single-day-large-sample",
                sheet_name="Sheet1",
                frame=frame,
                request=SmartReportRequest(sheet_name="Sheet1", use_r_workflow=True, core_purpose="single day robustness"),
                report_lens="generic_business_review",
            )

        workflow_dir = Path(result["workflow_dir"])
        self.assertTrue(workflow_dir.exists())
        method_log = pd.read_csv(workflow_dir / "method_log.csv")
        temporal_growth = method_log[method_log["method"] == "temporal_growth"]
        self.assertFalse(temporal_growth.empty)
        self.assertTrue((temporal_growth["status"].astype(str).str.lower() == "true").all())
        growth_csv = pd.read_csv(workflow_dir / "temporal_growth.csv")
        self.assertTrue(growth_csv.empty)

    def test_management_accounting_r_workflow_runs_all_numeric_fields_and_budget_variance(self) -> None:
        import shutil
        import uuid
        from unittest.mock import patch

        import app.services.r_workflow_service as rws
        import app.services.codex_service as cs
        from app.services.path_service import REPORTS_DIR
        import pandas as pd

        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-01-01", "2026-01-02", "2026-01-03", "2026-01-04"]),
                "责任中心": ["华东大区", "华东大区", "品牌中心", "品牌中心"],
                "产品线": ["珍护", "基础款", "珍护", "基础款"],
                "营业收入": [4000, 4100, 3900, 4200],
                "营业成本": [2600, 2620, 2550, 2680],
                "费用": [650, 660, 640, 670],
                "毛利": [1400, 1480, 1350, 1520],
                "净利润": [750, 820, 710, 850],
                "预算收入": [3950, 3980, 3920, 4010],
                "实际收入": [4000, 4100, 3900, 4200],
                "预算成本": [2580, 2590, 2520, 2600],
                "实际成本": [2600, 2620, 2550, 2680],
                "应收账款": [800, 830, 780, 860],
                "应付账款": [520, 540, 500, 560],
                "存货": [600, 620, 590, 650],
                "经营现金流": [420, 460, 390, 480],
                "总资产": [7800, 7900, 7700, 8050],
                "总负债": [3500, 3550, 3450, 3600],
                "所有者权益": [4300, 4350, 4250, 4450],
            }
        )

        report_id = "mgmtall" + uuid.uuid4().hex[:8]
        report_dir = REPORTS_DIR / f"smart-report-{report_id}"
        if report_dir.exists():
            shutil.rmtree(report_dir)

        with patch.object(rws, "codex_generate_r_workflow", lambda ctx: cs._fallback_r_workflow_author_v2("forced-test", ctx)), patch.object(
            rws, "codex_interpret_r_results", lambda payload: cs._fallback_r_results_interpretation("forced-test", payload)
        ), patch.object(
            rws, "codex_infer_r_workflow_semantics", lambda ctx: ctx["heuristic_registry"]
        ), patch.object(
            rws, "_resolve_rscript_path", return_value="fake-rscript"
        ), patch.object(
            rws, "_run_rscript", side_effect=_fake_rscript
        ), patch.object(
            rws, "_run_r_workflow_codex_sidecar", return_value={"enabled": False, "downloadables": []}
        ):
            result = run_r_workflow(
                report_dir=report_dir,
                report_id=report_id,
                dataset_name="management-accounting-sample",
                sheet_name="Sheet1",
                frame=frame,
                request=SmartReportRequest(sheet_name="Sheet1", use_r_workflow=True, core_purpose="management accounting"),
                report_lens="management_accounting_review",
            )

        workflow_dir = Path(result["workflow_dir"])
        category_metric = pd.read_csv(workflow_dir / "category_metric_summary.csv")
        metrics = sorted(category_metric["metric"].dropna().astype(str).unique().tolist())
        self.assertIn("经营现金流", metrics)
        self.assertIn("总资产", metrics)
        self.assertIn("所有者权益", metrics)
        self.assertIn("实际成本", metrics)
        self.assertGreaterEqual(len(metrics), 16)
        budget_variance = pd.read_csv(workflow_dir / "budget_variance_summary.csv")
        self.assertFalse(budget_variance.empty)
        self.assertIn("metric_pair", budget_variance.columns)
        self.assertIn("gap", budget_variance.columns)

    def test_single_day_r_workflow_filters_zero_variance_columns_before_pca_and_cluster(self) -> None:
        import shutil
        import uuid
        from unittest.mock import patch

        import app.services.codex_service as cs
        import app.services.r_workflow_service as rws
        import pandas as pd
        from app.services.path_service import REPORTS_DIR

        frame = pd.DataFrame(
            {
                "date": ["2017-12-02"] * 12,
                "item_id": [f"i{i}" for i in range(12)],
                "category_id": ["c1", "c2"] * 6,
                "page_views": [100, 120, 80, 95, 140, 160, 110, 130, 90, 105, 150, 170],
                "cart_count": [10, 12, 8, 9, 14, 16, 11, 13, 9, 10, 15, 17],
                "fav_count": [5, 6, 4, 4, 7, 8, 5, 6, 4, 5, 7, 8],
                "buy_count": [2, 3, 1, 2, 4, 5, 3, 3, 2, 2, 4, 5],
                "user_count": [1] * 12,
                "buy_rate": [1.0] * 12,
                "cart_rate": [1.0] * 12,
            }
        )

        report_id = "zerovar" + uuid.uuid4().hex[:8]
        report_dir = REPORTS_DIR / f"smart-report-{report_id}"
        if report_dir.exists():
            shutil.rmtree(report_dir)

        with patch.object(rws, "codex_generate_r_workflow", lambda ctx: cs._fallback_r_workflow_author_v2("forced-test", ctx)), patch.object(
            rws, "codex_interpret_r_results", lambda payload: cs._fallback_r_results_interpretation("forced-test", payload)
        ), patch.object(
            rws, "codex_infer_r_workflow_semantics", lambda ctx: ctx["heuristic_registry"]
        ), patch.object(
            rws, "_resolve_rscript_path", return_value="fake-rscript"
        ), patch.object(
            rws, "_run_rscript", side_effect=_fake_rscript
        ), patch.object(
            rws, "_run_r_workflow_codex_sidecar", return_value={"enabled": False, "downloadables": []}
        ):
            result = run_r_workflow(
                report_dir=report_dir,
                report_id=report_id,
                dataset_name="single-day-zero-variance",
                sheet_name="Sheet1",
                frame=frame,
                request=SmartReportRequest(sheet_name="Sheet1", use_r_workflow=True, core_purpose="zero variance robustness"),
                report_lens="generic_business_review",
            )

        workflow_dir = Path(result["workflow_dir"])
        method_log = pd.read_csv(workflow_dir / "method_log.csv")
        pca_variance = method_log[method_log["method"] == "pca_variance"].iloc[0]
        kmeans_clusters = method_log[method_log["method"] == "kmeans_clusters"].iloc[0]
        self.assertEqual(str(pca_variance["status"]).lower(), "true")
        self.assertEqual(str(kmeans_clusters["status"]).lower(), "true")
        self.assertTrue((workflow_dir / "pca_axis_summary.csv").exists())
        self.assertTrue((workflow_dir / "cluster_member_detail.csv").exists())
        loadings = pd.read_csv(workflow_dir / "pca_loadings.csv")
        variables = loadings["variable"].astype(str).tolist()
        self.assertNotIn("user_count", variables)
        self.assertNotIn("buy_rate", variables)
        self.assertNotIn("cart_rate", variables)
        cluster_detail = pd.read_csv(workflow_dir / "cluster_member_detail.csv")
        self.assertIn("cluster", cluster_detail.columns.astype(str).tolist())
        self.assertIn("row_index", cluster_detail.columns.astype(str).tolist())


if __name__ == "__main__":
    unittest.main()
