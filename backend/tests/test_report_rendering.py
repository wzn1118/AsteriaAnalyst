from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import pandas as pd

from app.models import SmartReportRequest
from app.services.management_accounting_knowledge_service import build_management_accounting_context
from app.services.path_service import REPORTS_DIR
from app.services.internet_operations_pdf_report_renderer import build_internet_operations_management_variant
from app.services.report_service import (
    _expert_recall_bullets,
    _downloadable_bundle_cn,
    _procurement_full_appendix_variant,
    _extend_report_with_semantic_metric_sections,
    _generic_management_sections_from_program,
    _business_background_sections_from_codex,
    _commercial_dimension_sections_from_codex,
    _business_object_sections_from_codex,
    _infer_report_lens,
    _internet_ops_sections_from_agent,
    _management_section_ids,
    _metric_interpretation_section,
    _management_accounting_sections_from_context,
    _merge_multi_sheet_frames,
    _media_resource_bullets,
    _ops_activity_scorecard_section,
    _ops_channel_scorecard_section,
    _ops_combo_direct_actions_section,
    _ops_combo_impact_section,
    _ensure_internet_ops_impact_sections,
    _ops_dimension_effect_section,
    _ops_content_scorecard_section,
    _ops_topline_section,
    _render_report_html_cn,
    _report_markdown_cn,
    _resolved_multi_sheet_names,
    _statistical_numeric_columns,
)


class ReportRenderingTests(unittest.TestCase):
    def _sample_report(self) -> dict:
        return {
            "title": "测试报告",
            "dataset_name": "测试数据集",
            "sheet_name": "Sheet1",
            "report_language": "zh-CN",
            "generated_at": "2026-04-16T00:00:00Z",
            "executive_summary": ["一句话总判断", "高风险预警", "优先动作A"],
            "requirement_restatement": {"bullets": ["需求1"]},
            "requirement_confirmation": {"items": ["假设1"]},
            "data_initial_understanding": {"bullets": ["理解1"]},
            "output_strategy": {"bullets": ["主输出1"]},
            "generation_plan": ["计划1"],
            "historical_report_adaptation": None,
            "tool_integrations": [],
            "sections": [
                {
                    "id": "management_story",
                    "title": "关键发现展开",
                    "summary": "管理层正文摘要",
                    "bullets": ["发现1"],
                    "tables": [{"title": "动作矩阵", "columns": ["对象", "动作"], "rows": [{"对象": "A", "动作": "优先排查"}]}],
                    "charts": [],
                },
                {
                    "id": "numeric",
                    "title": "数值画像",
                    "summary": "附录摘要",
                    "bullets": ["附录1"],
                    "tables": [{"title": "表1", "columns": ["字段", "值"], "rows": [{"字段": "CTR", "值": "2.6%"}]}],
                    "charts": [],
                },
                {
                    "id": "field_dictionary",
                    "title": "字段词典",
                    "summary": "附录页",
                    "bullets": ["字段说明"],
                    "tables": [{"title": "字段表", "columns": ["字段", "说明"], "rows": [{"字段": "CTR", "说明": "点击率"}]}],
                    "charts": [],
                },
            ],
        }

    def test_markdown_splits_management_and_appendix(self) -> None:
        markdown = _report_markdown_cn(self._sample_report())
        self.assertIn("## 管理层正文", markdown)
        self.assertIn("## 分析附录", markdown)
        self.assertIn("### 动作矩阵", markdown)

    def test_procurement_full_appendix_variant_keeps_full_sections_and_raw_appendix(self) -> None:
        report = self._sample_report()
        report["report_lens"] = "procurement_sales_review"
        report["report_id"] = "demo-procurement-report"
        report["object_decision_registry"] = {
            "rows": [
                {
                    "object_level": "category",
                    "object_id": "health_beauty",
                    "final_label": "头部品类资源候选",
                    "final_action": "待补毛利后判断是否放量",
                    "conclusion_type": "proxy_based_inference",
                    "confidence_level": "medium",
                    "blocked_actions": ["资源倾斜"],
                    "validation_plan": "补毛利后复核",
                }
            ],
            "blocked_action_log": [
                {
                    "object": "health_beauty",
                    "object_level": "category",
                    "attempted_action": "资源倾斜",
                    "blocked_reason": "缺利润字段",
                    "downgraded_action": "人工复核",
                    "missing_fields": "profit_fields",
                }
            ],
            "data_gap_rows": [{"字段域": "profit_fields", "是否可用": "否", "命中字段": "无"}],
            "field_registry": {
                "field_coverage": {"sales_fields": ["Revenue"], "profit_fields": []},
                "has_sales_fields": True,
                "has_fulfillment_fields": True,
                "has_review_fields": True,
                "has_profit_fields": False,
                "has_inventory_fields": False,
                "has_procurement_price_fields": False,
                "has_supplier_delivery_fields": False,
                "has_payment_term_fields": False,
                "has_return_cost_fields": False,
                "has_supplier_contract_fields": False,
                "has_competitor_price_fields": False,
            },
        }

        variant = _procurement_full_appendix_variant(
            report,
            report["object_decision_registry"],
            report["object_decision_registry"]["field_registry"],
        )

        section_ids = [section["id"] for section in variant["sections"]]
        self.assertIn("management_story", section_ids)
        self.assertIn("numeric", section_ids)
        self.assertIn("appendix_registry", section_ids)
        self.assertIn("appendix_blocked", section_ids)
        self.assertIn("appendix_field_registry", section_ids)

    def test_html_bundle_renders_tables(self) -> None:
        with tempfile.TemporaryDirectory(dir=str(REPORTS_DIR.parent)) as tmpdir:
            report_dir = Path(tmpdir)
            frame = pd.DataFrame({"A": [1]})
            bundle = _downloadable_bundle_cn(report_dir, "demo", self._sample_report(), frame, {})
            html_path = Path(bundle["html_path"])
            html_text = html_path.read_text(encoding="utf-8")
            self.assertIn("<table", html_text)
            self.assertIn("table-scroll", html_text)
            self.assertIn("管理层正文", html_text)
            self.assertIn("分析附录", html_text)

    def test_management_sections_exclude_engineering_and_thinking_chain(self) -> None:
        section_ids = _management_section_ids()
        self.assertNotIn("engineering_chain", section_ids)
        self.assertNotIn("thinking_chain", section_ids)

    def test_media_resource_bullets_stay_conservative_when_shares_are_balanced(self) -> None:
        bullets = _media_resource_bullets(
            {
                "dimension_column": "渠道",
                "segment_rows": [
                    {"segment": "社群", "share": 0.169},
                    {"segment": "自然", "share": 0.168},
                    {"segment": "搜索", "share": 0.166},
                ],
                "top3_share": 0.503,
                "hhi": 1666.7,
            }
        )
        joined = " ".join(bullets)
        self.assertIn("不适合直接写成头部格局", joined)
        self.assertNotIn("第一优先对象", joined)

    def test_infer_report_lens_prefers_primary_family_over_budget_keyword(self) -> None:
        request = SmartReportRequest(
            user_requirement="请输出一份面向财务和经营负责人的管理会计分析报告",
            problem_to_solve="判断利润质量、预算偏差、资金占用和责任中心差异",
            target_audience="财务负责人 / 经营负责人 / 管理层",
            core_purpose="形成财务经营联动复盘与资源配置判断",
            expected_result="一份中文主报告",
        )
        lens = _infer_report_lens(
            request,
            "经营管理会计样本",
            {
                "task_model": {"primary_family": "management_accounting_review"},
                "object_candidates": [{"object_type": "management_accounting_statement"}],
                "lens": "management_accounting_review",
            },
        )
        self.assertEqual(lens, "management_accounting_review")

    def test_infer_report_lens_honors_explicit_business_profile(self) -> None:
        request = SmartReportRequest(
            business_profile="internet_operations_report",
            user_requirement="做互联网运营复盘",
        )
        lens = _infer_report_lens(
            request,
            "eval-ops-100000-r1",
            {
                "task_model": {"primary_family": "mixed_business_review"},
                "object_candidates": [{"object_type": "generic_business_table"}],
                "lens": "mixed_business_review",
            },
        )
        self.assertEqual(lens, "internet_ops_review")


    def test_excel_bundle_still_writes_when_no_tables_exist(self) -> None:
        report = self._sample_report()
        report["sections"] = [
            {
                "id": "management_story",
                "title": "empty",
                "summary": "no tables",
                "bullets": ["text only"],
                "tables": [],
                "charts": [],
            }
        ]
        with tempfile.TemporaryDirectory(dir=str(REPORTS_DIR.parent)) as tmpdir:
            report_dir = Path(tmpdir)
            frame = pd.DataFrame({"A": [1]})
            bundle = _downloadable_bundle_cn(report_dir, "demo-empty", report, frame, {})
            self.assertTrue(Path(bundle["excel_path"]).exists())

    def test_excel_bundle_keeps_all_section_tables(self) -> None:
        try:
            from openpyxl import load_workbook
        except Exception:
            self.skipTest("openpyxl unavailable")

        report = self._sample_report()
        report["sections"][1]["tables"].append(
            {"title": "第二张附录表", "columns": ["字段", "值"], "rows": [{"字段": "CVR", "值": "1.8%"}]}
        )
        report["sections"][1]["tables"].append(
            {"title": "第三张附录表", "columns": ["字段", "值"], "rows": [{"字段": "ARPU", "值": "9.2"}]}
        )
        with tempfile.TemporaryDirectory(dir=str(REPORTS_DIR.parent)) as tmpdir:
            report_dir = Path(tmpdir)
            frame = pd.DataFrame({"A": [1]})
            bundle = _downloadable_bundle_cn(report_dir, "demo-alltables", report, frame, {})
            workbook = load_workbook(bundle["excel_path"])
            self.assertGreaterEqual(len(workbook.sheetnames), 4)

    def test_semantic_metric_artifacts_bind_to_report_and_excel(self) -> None:
        try:
            from openpyxl import load_workbook
        except Exception:
            self.skipTest("openpyxl unavailable")

        with tempfile.TemporaryDirectory(dir=str(REPORTS_DIR.parent)) as tmpdir:
            report_dir = Path(tmpdir)
            metric_dir = report_dir / "outputs" / "metric_mining"
            metric_dir.mkdir(parents=True, exist_ok=True)
            semantic_result_path = metric_dir / "semantic_metric_result.json"
            metric_log_path = metric_dir / "metric_derivation_log.csv"
            semantic_result = {
                "trace_id": "trace-plan-001",
                "results": [
                    {
                        "metric_id": "gross_margin",
                        "metric_name_cn": "毛利率",
                        "value": 0.32,
                        "grain": "overall",
                        "source_fields": ["revenue", "cost"],
                        "formula": "(revenue - cost) / revenue",
                        "calculation_method": "pandas deterministic executor",
                        "evidence_level": "B_DERIVED",
                        "confidence": 0.91,
                        "status": "calculated",
                    }
                ],
            }
            semantic_result_path.write_text(json.dumps(semantic_result, ensure_ascii=False), encoding="utf-8")
            (metric_dir / "semantic_metric_plan.json").write_text(
                json.dumps(
                    {
                        "metric_plans": [
                            {
                                "metric_id": "gross_margin",
                                "metric_name_cn": "毛利率",
                                "metric_family": "profitability_metric",
                                "business_object": "overall",
                                "metric_type": "derived",
                                "matched_fields": ["revenue", "cost"],
                                "missing_fields": [],
                                "formula_or_logic": "(revenue - cost) / revenue",
                                "grain": "overall",
                                "evidence_level": "B_DERIVED",
                                "confidence": 0.91,
                                "allowed_downstream_usage": "profitability context",
                                "forbidden_downstream_usage": "SKU-level profit claim",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (metric_dir / "metric_opportunity_graph.json").write_text(
                json.dumps(
                    {
                        "nodes": [
                            {
                                "metric_family": "profitability_metric",
                                "candidate_metric_ids": ["gross_margin"],
                                "best_feasibility": "calculable",
                                "matched_patterns": [{"matched_fields": ["revenue", "cost"]}],
                                "forbidden_without_roles": ["cost"],
                                "calculation_executor": "deterministic_metric_executor",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            metric_log_path.write_text(
                "metric_id,status,source_fields\n"
                "gross_margin,calculated,\"revenue,cost\"\n",
                encoding="utf-8",
            )

            report = self._sample_report()
            report["semantic_metric_result"] = semantic_result
            report["semantic_metric_result_path"] = str(semantic_result_path)
            report["metric_derivation_log_path"] = str(metric_log_path)
            _extend_report_with_semantic_metric_sections(report)

            section_ids = {section["id"] for section in report["sections"]}
            self.assertIn("semantic_metric_scorecard", section_ids)
            self.assertIn("appendix_semantic_metric_result", section_ids)
            self.assertIn("appendix_semantic_metric_plan", section_ids)
            self.assertIn("appendix_metric_derivation_log", section_ids)

            frame = pd.DataFrame({"A": [1]})
            bundle = _downloadable_bundle_cn(report_dir, "demo-semantic", report, frame, {})
            workbook = load_workbook(bundle["excel_path"])
            sheet_joined = " ".join(workbook.sheetnames)
            self.assertIn("semantic_metric", sheet_joined)
            self.assertIn("appendix_s", sheet_joined)

    def test_semantic_metric_appendix_localizes_headers_and_formats_multiline_cells(self) -> None:
        report = self._sample_report()
        report["semantic_metric_result"] = {
            "results": [
                {
                    "metric_id": "category_revenue_share",
                    "metric_name_cn": "品类营收占比",
                    "value": [
                        {"dimension": "health_beauty", "value": 10168.18, "share": 0.2004},
                        {"dimension": "housewares", "value": 7800.22, "share": 0.1537},
                    ],
                    "grain": "dataset total; can also GROUP BY Category",
                    "source_fields": ["Revenue", "Category"],
                    "formula": "SUM(Revenue) by Category / SUM(Revenue) overall",
                    "calculation_method": "deterministic_share_executor",
                    "evidence_level": "B_DERIVED",
                    "confidence": 0.95,
                    "caveat": "缺成本时不能把结构占比写成利润结论。",
                    "status": "calculated",
                }
            ]
        }
        report["semantic_metric_result_path"] = "fake/semantic_metric_result.json"
        _extend_report_with_semantic_metric_sections(report)

        appendix_section = next(
            section for section in report["sections"] if section["id"] == "appendix_semantic_metric_result"
        )
        table = appendix_section["tables"][0]
        self.assertIn("来源字段", table["columns"])
        self.assertIn("计算方法", table["columns"])
        first_row = table["rows"][0]
        self.assertIn("\n", first_row["来源字段"])
        self.assertIn("\n", first_row["数值"])

        markdown = _report_markdown_cn(report)
        html = _render_report_html_cn(report)
        self.assertIn("来源字段", markdown)
        self.assertIn("计算方法", markdown)
        self.assertIn("<br>", markdown)
        self.assertIn("<br/>", html)
        self.assertIn("智能指标结构化明细附录", markdown)
        self.assertIn("semantic_metric_result_expanded", markdown)

    def test_generic_management_sections_expand_report_body(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-04-01", "2026-04-02", "2026-04-03"]),
                "渠道": ["自然", "广告", "社群"],
                "活跃用户": [1200, 980, 1100],
                "注册数": [120, 88, 105],
                "留存率": [0.42, 0.36, 0.4],
            }
        )
        sections = _generic_management_sections_from_program(
            report_lens="internet_ops_review",
            request=SmartReportRequest(
                user_requirement="生成互联网运营复盘",
                problem_to_solve="找到新增和留存的主要差异",
                target_audience="运营负责人 / 增长负责人",
                core_purpose="形成周会复盘与后续动作",
                expected_result="主报告 + 后续动作清单",
            ),
            program_bundle={
                "task_model": {"problem": "找到新增和留存差异", "purpose": "形成增长判断"},
                "program": {
                    "observation_unit": "单日渠道记录",
                    "core_outcomes": ["活跃用户", "注册数"],
                    "efficiency_metrics": ["留存率"],
                    "explanatory_slices": ["渠道", "日期"],
                    "can_analyze": ["新增与留存差异", "渠道结构"],
                    "cannot_analyze": ["真实用户生命周期价值"],
                    "confidence": "high",
                },
                "object_candidates": [{"object_type": "internet_operations_log"}],
                "ai_hints": {
                    "task_family_candidates": [{"family": "internet_ops_review", "why": "字段更像运营复盘"}],
                    "object_candidates": [{"object_type": "internet_operations_log", "observation_unit": "单日渠道记录"}],
                },
                "experts": [
                    {
                        "expert": "互联网运营分析负责人",
                        "focus": "优先看新增、活跃和留存差异",
                        "priority_questions": ["新增从哪里来", "留存为什么掉"],
                        "decision_outputs": ["增长动作", "复盘结论"],
                    }
                ],
                "hypotheses": [{"title": "internet_operations_log", "confidence": "high", "why": "字段组合完整"}],
                "derived_metrics": [{"metric": "留存率", "formula": "留存用户 / 新增用户", "confidence": "high"}],
            },
            frame=frame,
            market_intelligence={"ready": False},
            temporal_rows=[
                {"period": "2026-04-01", "value": 1200},
                {"period": "2026-04-02", "value": 980},
                {"period": "2026-04-03", "value": 1100},
            ],
            correlation_rows=[{"left": "活跃用户", "right": "注册数", "correlation": 0.91}],
            outlier_rows=[{"column": "注册数", "outlier_ratio": 0.12}],
        )
        section_ids = {section["id"] for section in sections}
        self.assertTrue({"input_alignment", "scale_snapshot", "expert_panel", "hypothesis_board", "engineering_chain", "thinking_chain", "metric_contract", "evidence_chain", "industry_playbook", "role_guide", "management_story", "meeting_ready", "boss_qa", "execution_implication", "scenario_matrix", "decision_guardrails", "learning_agenda", "action_roadmap"} <= section_ids)
        metric_section = next(section for section in sections if section["id"] == "metric_contract")
        self.assertTrue(metric_section["tables"])
        meeting_section = next(section for section in sections if section["id"] == "meeting_ready")
        first_row = meeting_section["tables"][0]["rows"][0]
        self.assertIn("internet_operations_log", first_row["要讲什么"])
        self.assertIn("Rows=3", meeting_section["tables"][0]["rows"][1]["建议展示"])
        self.assertIn("活跃用户", meeting_section["tables"][0]["rows"][1]["要讲什么"])

    def test_ops_management_sections_include_business_tables(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-04-01", "2026-04-01", "2026-04-02", "2026-04-02"]),
                "渠道": ["自然", "广告", "自然", "社群"],
                "活动": ["拉新活动", "拉新活动", "召回活动", "内容上新"],
                "内容主题": ["品牌故事", "产品卖点", "品牌故事", "策略教程"],
                "活跃用户": [1200, 980, 1400, 600],
                "新增用户": [120, 88, 180, 60],
                "留存用户": [500, 360, 630, 250],
                "留存率": [0.42, 0.36, 0.45, 0.40],
                "转化率": [0.11, 0.08, 0.12, 0.09],
                "订单数": [40, 28, 55, 19],
            }
        )
        self.assertIsNotNone(_ops_topline_section(frame))
        self.assertIsNotNone(_ops_channel_scorecard_section(frame))
        self.assertIsNotNone(_ops_activity_scorecard_section(frame))
        self.assertIsNotNone(_ops_content_scorecard_section(frame))
        self.assertIsNotNone(_ops_dimension_effect_section("test", frame, dimension_column="渠道", label="渠道", section_id="ops_channel_impact"))
        self.assertIsNotNone(_ops_dimension_effect_section("test", frame, dimension_column="活动", label="活动", section_id="ops_activity_impact"))
        self.assertIsNotNone(_ops_dimension_effect_section("test", frame, dimension_column="内容主题", label="内容主题", section_id="ops_content_impact"))
        self.assertIsNotNone(_ops_combo_impact_section("test", frame))
        self.assertIsNotNone(_ops_combo_direct_actions_section("test", frame))

    def test_ops_combo_direct_actions_are_not_meta_table_building(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-04-01"] * 6 + ["2026-04-02"] * 6),
                "渠道": ["联盟", "联盟", "自然", "自然", "广告", "广告"] * 2,
                "活动": ["召回", "拉新", "会员促活", "内容上新", "召回", "内容上新"] * 2,
                "内容主题": ["策略教程", "品牌故事", "策略教程", "品牌故事", "产品卖点", "产品卖点"] * 2,
                "活跃用户": [900, 300, 520, 410, 460, 430, 920, 320, 540, 420, 455, 440],
                "新增用户": [180, 40, 60, 55, 80, 70, 185, 42, 58, 54, 82, 68],
                "留存用户": [430, 90, 180, 150, 155, 150, 440, 92, 182, 151, 150, 152],
                "留存率": [0.48, 0.30, 0.34, 0.36, 0.33, 0.35, 0.48, 0.29, 0.34, 0.36, 0.33, 0.35],
                "转化率": [0.13, 0.05, 0.08, 0.09, 0.07, 0.08, 0.13, 0.05, 0.08, 0.09, 0.07, 0.08],
                "订单数": [72, 11, 25, 24, 20, 22, 74, 12, 24, 23, 19, 22],
            }
        )
        section = _ops_combo_direct_actions_section("test", frame)
        self.assertIsNotNone(section)
        table = section["tables"][0]
        first_row = table["rows"][0]
        joined = " ".join(str(value) for value in first_row.values())
        self.assertNotIn("建表", joined)
        self.assertNotIn("看板", joined)
        self.assertIn("直接", joined)

    def test_ops_sections_do_not_force_priority_when_differences_are_close(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-04-01"] * 6),
                "渠道": ["社群", "自然", "广告", "Push", "联盟", "搜索"],
                "活动": ["内容上新", "会员促活", "搜索冲量", "召回活动", "拉新活动", "内容上新"],
                "内容主题": ["品牌故事", "品牌故事", "产品卖点", "策略教程", "福利活动", "产品卖点"],
                "活跃用户": [1000, 995, 990, 998, 992, 989],
                "新增用户": [120, 119, 118, 121, 120, 119],
                "留存用户": [420, 418, 417, 421, 419, 418],
                "留存率": [0.420, 0.421, 0.419, 0.420, 0.421, 0.420],
                "转化率": [0.050, 0.0505, 0.0498, 0.0501, 0.0500, 0.0499],
                "订单数": [52, 51, 50, 51, 50, 50],
            }
        )
        channel_section = _ops_channel_scorecard_section(frame)
        activity_section = _ops_activity_scorecard_section(frame)
        content_section = _ops_content_scorecard_section(frame)
        joined = " ".join(channel_section["bullets"] + activity_section["bullets"] + content_section["bullets"])
        self.assertIn("不能直接", joined)
        self.assertIn("样本量最大不等于活动机制最好", joined)

    def test_ops_dimension_effect_section_compares_total_and_efficiency(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-04-01"] * 6),
                "渠道": ["社群", "自然", "广告", "Push", "联盟", "搜索"],
                "活动": ["内容上新", "会员促活", "搜索冲量", "召回活动", "拉新活动", "内容上新"],
                "内容主题": ["品牌故事", "品牌故事", "产品卖点", "策略教程", "福利活动", "产品卖点"],
                "活跃用户": [1000, 995, 990, 998, 992, 989],
                "新增用户": [120, 119, 118, 121, 120, 119],
                "留存用户": [420, 418, 417, 421, 419, 418],
                "留存率": [0.420, 0.421, 0.419, 0.420, 0.421, 0.420],
                "转化率": [0.050, 0.0505, 0.0498, 0.0501, 0.0500, 0.0499],
                "订单数": [52, 51, 50, 51, 50, 50],
            }
        )
        section = _ops_dimension_effect_section("test", frame, dimension_column="渠道", label="渠道", section_id="ops_channel_impact")
        self.assertIsNotNone(section)
        self.assertIn("直接评估不同渠道对用户结果的实际影响", section["summary"])
        table = section["tables"][0]
        self.assertIn("活跃均值相对整体", table["columns"])
        self.assertIn("订单均值相对整体", table["columns"])

    def test_ops_combo_impact_section_evaluates_actual_combinations(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(
                    [
                        "2026-04-01",
                        "2026-04-02",
                        "2026-04-03",
                        "2026-04-04",
                        "2026-04-01",
                        "2026-04-02",
                        "2026-04-03",
                        "2026-04-04",
                    ]
                ),
                "渠道": ["社群", "社群", "自然", "自然", "Push", "Push", "广告", "广告"],
                "活动": ["内容上新", "内容上新", "内容上新", "内容上新", "召回活动", "召回活动", "搜索冲量", "搜索冲量"],
                "内容主题": ["攻略教程", "攻略教程", "品牌故事", "品牌故事", "用户案例", "用户案例", "品牌故事", "品牌故事"],
                "活跃用户": [2000, 2100, 1800, 1820, 1500, 1490, 2200, 2210],
                "新增用户": [380, 390, 310, 305, 260, 255, 330, 320],
                "留存用户": [980, 1005, 760, 770, 540, 535, 820, 825],
                "留存率": [0.49, 0.479, 0.422, 0.423, 0.36, 0.359, 0.373, 0.373],
                "转化率": [0.072, 0.071, 0.043, 0.044, 0.031, 0.03, 0.081, 0.08],
                "订单数": [145, 149, 92, 90, 61, 60, 188, 190],
            }
        )
        section = _ops_combo_impact_section("test", frame)
        self.assertIsNotNone(section)
        self.assertEqual(section["id"], "ops_combo_impact")
        self.assertIn("组合对新增、留存、转化和订单承接的实际影响", section["summary"])
        table = section["tables"][0]
        self.assertIn("组合角色", table["columns"])
        self.assertIn("稳定性", table["columns"])
        self.assertIn("订单均值相对整体", table["columns"])
        joined = " ".join(section["bullets"])
        self.assertIn("总量头部", joined)
        self.assertIn("订单均值", joined)

    def test_ensure_internet_ops_impact_sections_adds_missing_sections(self) -> None:
        frame = pd.DataFrame(
            {
                "日期": pd.to_datetime(["2026-04-01", "2026-04-02", "2026-04-03", "2026-04-04"]),
                "渠道": ["社群", "自然", "广告", "Push"],
                "活动": ["内容上新", "召回活动", "搜索冲量", "会员促活"],
                "内容主题": ["攻略教程", "品牌故事", "福利活动", "用户案例"],
                "活跃用户": [2000, 1800, 2200, 1500],
                "新增用户": [380, 310, 330, 260],
                "留存用户": [980, 760, 820, 540],
                "留存率": [0.49, 0.422, 0.373, 0.36],
                "转化率": [0.072, 0.043, 0.081, 0.031],
                "订单数": [145, 92, 188, 61],
            }
        )
        sections = _ensure_internet_ops_impact_sections(
            [{"id": "ops_specialist_judgement", "title": "判断", "summary": "s", "bullets": ["b"], "tables": [], "charts": []}],
            dataset_id="test",
            frame=frame,
            field_registry={"has_channel_fields": True, "has_campaign_fields": True, "has_content_fields": True},
        )
        ids = [section["id"] for section in sections]
        self.assertIn("ops_channel_impact", ids)
        self.assertIn("ops_activity_impact", ids)
        self.assertIn("ops_content_impact", ids)
        self.assertIn("ops_combo_impact", ids)
        self.assertIn("ops_combo_direct_actions", ids)

    def test_internet_ops_management_variant_uses_core_sections(self) -> None:
        report = {
            "title": "互联网运营报告",
            "business_profile": "internet_operations_report",
            "internet_ops_field_availability_registry": {
                "report_mode": "internet_operations_report",
                "available_field_groups": ["channel_fields", "campaign_fields", "content_fields", "user_fields", "retention_fields"],
                "missing_field_groups": [],
                "supported_analysis_modules": ["channel_operations", "campaign_review", "content_operations"],
                "unsupported_analysis_modules": [],
                "matched_field_signals": {},
            },
            "internet_operations_analysis_modules": {
                "north_star_metric_selector": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "aarrr_funnel_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "traffic_structure_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "user_growth_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "funnel_conversion_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "retention_cohort_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "content_operations_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "content_asset_matrix": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "channel_operations_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "campaign_operations_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "community_operations_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "monetization_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
                "risk_and_anomaly_analyzer": {"business_question": "q", "core_conclusion": "c", "evidence": {}, "missing_fields": [], "recommended_validation_action": "a", "validation_metric": "m"},
            },
            "internet_ops_object_decision_registry": {"rows": []},
            "internet_ops_action_table": [],
            "internet_ops_action_roadmap": {},
            "internet_ops_management_core_sections": [
                {"id": "ops_channel_impact", "title": "渠道影响评估", "summary": "s", "bullets": ["b"], "tables": [], "charts": []},
                {"id": "ops_combo_impact", "title": "组合影响评估", "summary": "s", "bullets": ["b"], "tables": [], "charts": []},
                {"id": "ops_combo_direct_actions", "title": "组合直接动作", "summary": "s", "bullets": ["b"], "tables": [], "charts": []},
            ],
        }
        variant = build_internet_operations_management_variant(report)
        self.assertIsNotNone(variant)
        ids = [section["id"] for section in variant["sections"]]
        self.assertIn("ops_channel_impact", ids)
        self.assertIn("ops_combo_impact", ids)
        self.assertIn("ops_combo_direct_actions", ids)

    def test_business_background_sections_render_from_codex(self) -> None:
        sections = _business_background_sections_from_codex(
            {
                "background_summary": "这份背景说明当前分析要服务于渠道和活动资源分配。",
                "key_points": ["社群要看承接质量", "内容上新要区分冲量和稳定承接"],
                "decision_implications": ["不能只按活跃总量排优先级"],
            }
        )
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0]["id"], "business_background_deep_read")
        self.assertTrue(sections[0]["tables"])

    def test_business_object_sections_render_from_codex(self) -> None:
        sections = _business_object_sections_from_codex(
            {
                "background_summary": "系统已经自动判断头部对象在业务链里的角色。",
                "key_points": ["社群更像私域承接渠道"],
                "object_meanings": [
                    {
                        "dimension": "渠道",
                        "value": "社群",
                        "inferred_role": "私域承接渠道",
                        "business_meaning": "社群更适合看承接质量。",
                        "action_hint": "先比较留存和订单承接。",
                    }
                ],
            }
        )
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0]["id"], "business_object_meaning")
        self.assertTrue(sections[0]["tables"])

    def test_commercial_dimension_sections_render_from_codex(self) -> None:
        sections = _commercial_dimension_sections_from_codex(
            {
                "management_summary": "这份数据适合按京东采销口径来读。",
                "priority_actions": ["把 SKU 按引流款、承接款、利润款拆开管理。"],
                "dimension_reviews": [
                    {
                        "dimension": "SKU",
                        "headline": "SKU 当前头部对象是 `六神止痒花露水`",
                        "finding": "它的总量更高，但还要看利润和库存。",
                        "business_action": "继续比较主推与低效 SKU 的角色差异。",
                    }
                ],
            }
        )
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0]["id"], "commercial_dimension_judgement")
        self.assertTrue(sections[0]["tables"])

    def test_statistical_numeric_columns_excludes_year_like_and_constant_fields(self) -> None:
        frame = pd.DataFrame(
            {
                "年度": [2024, 2024, 2024, 2024],
                "本年度总支出": [100.0, 200.0, 300.0, 400.0],
                "本年度收入合计": [120.0, 210.0, 350.0, 420.0],
            }
        )
        filtered = _statistical_numeric_columns(frame, ["年度", "本年度总支出", "本年度收入合计"], ["年度"])
        self.assertEqual(filtered, ["本年度总支出", "本年度收入合计"])

    def test_management_accounting_sections_render_from_context(self) -> None:
        frame = pd.DataFrame(
            {
                "期间": ["2026Q1", "2026Q1", "2026Q2", "2026Q2"],
                "责任中心": ["华东", "华南", "华东", "华南"],
                "营业收入": [1000.0, 800.0, 1100.0, 780.0],
                "营业成本": [620.0, 520.0, 670.0, 540.0],
                "净利润": [120.0, 70.0, 150.0, 40.0],
                "预算收入": [950.0, 820.0, 1080.0, 800.0],
                "实际收入": [1000.0, 800.0, 1100.0, 780.0],
                "应收账款": [180.0, 150.0, 200.0, 165.0],
                "应付账款": [120.0, 110.0, 130.0, 118.0],
                "存货": [160.0, 120.0, 170.0, 128.0],
            }
        )
        context = build_management_accounting_context(
            frame=frame,
            request=SmartReportRequest(
                user_requirement="请生成管理会计分析报告",
                problem_to_solve="判断利润质量、预算偏差和营运资本压力",
                target_audience="财务负责人 / 经营负责人",
                core_purpose="形成管理复盘",
                expected_result="主报告 + 风险清单",
            ),
            dataset_name="经营管理会计样本",
            program_bundle={"task_model": {"primary_family": "management_accounting_review"}},
        )
        self.assertIsNotNone(context)
        context["agent_layer"] = {
            "management_summary": "当前样本更适合按管理会计口径做利润、预算、现金和责任中心复盘。",
            "key_findings": ["净利率可直接用于区分规模增长和质量增长。"],
            "metric_interpretations": ["预算偏差要先拆价格、销量和效率来源。"],
            "risk_flags": ["营运资本占用不能被利润增长掩盖。"],
            "managerial_actions": ["优先复盘偏差最大的预算主题。"],
        }
        sections = _management_accounting_sections_from_context(context)
        section_ids = {section["id"] for section in sections}
        self.assertTrue({"management_accounting_profitability", "management_accounting_budget", "management_accounting_capital", "management_accounting_responsibility", "management_accounting_playbook"} <= section_ids)

    def test_internet_ops_agent_sections_render(self) -> None:
        sections = _internet_ops_sections_from_agent(
            {
                "management_summary": "这份数据更适合按互联网运营复盘口径来组织。",
                "key_findings": ["社群是当前最值得优先复盘的渠道。"],
                "metric_interpretations": ["新增用户需要结合留存率一起看。"],
                "risk_flags": ["高表现活动需要确认是否由单次峰值驱动。"],
                "managerial_actions": ["先围绕社群做渠道深挖。"],
                "validation_agenda": ["下一轮验证应拆到日期窗口。"],
            }
        )
        section_ids = {section["id"] for section in sections}
        self.assertEqual(section_ids, {"ops_specialist_judgement", "ops_specialist_actions"})
        self.assertIn("互联网运营专门结论层", sections[0]["title"])

    def test_resolved_multi_sheet_names_prefers_explicit_selection(self) -> None:
        metadata = {
            "active_sheet": "Sheet1",
            "sheets": [
                {"name": "Sheet1"},
                {"name": "Sheet2"},
                {"name": "Sheet3"},
            ],
        }
        request = SmartReportRequest(
            selected_sheets=["Sheet2", "Sheet3", "Missing"],
            multi_table_mode="combined",
        )
        self.assertEqual(_resolved_multi_sheet_names(metadata, request), ["Sheet2", "Sheet3"])

    def test_merge_multi_sheet_frames_adds_source_sheet_column(self) -> None:
        metadata = {"name": "Demo Workbook"}
        frame_a = pd.DataFrame({"指标": ["A"], "数值": [1]})
        frame_b = pd.DataFrame({"指标": ["B"], "金额": [2]})
        merged, synthetic_sheet = _merge_multi_sheet_frames(
            metadata,
            [
                ({"name": "Sheet1"}, frame_a),
                ({"name": "Sheet2"}, frame_b),
            ],
        )
        self.assertEqual(synthetic_sheet["name"], "合并分析(Sheet1+Sheet2)")
        self.assertIn("来源工作表", merged.columns)
        self.assertEqual(set(merged["来源工作表"].tolist()), {"Sheet1", "Sheet2"})

    def test_expert_recall_bullets_include_soul_language(self) -> None:
        bullets = _expert_recall_bullets(
            {
                "experts": [
                    {
                        "expert": "管理会计经营分析负责人",
                        "focus": "优先判断利润质量和预算偏差",
                        "priority_questions": ["利润是否有质量"],
                        "evidence_preference": ["收入与利润指标"],
                        "decision_outputs": ["经营质量复盘"],
                        "guardrails": ["不把收入增长直接写成经营改善"],
                        "obsession": "收入不是终点，利润质量才是盘子。",
                        "signature_move": "先看规模和利润率，再拆预算偏差。",
                        "anti_pattern": "把收入增长直接写成经营改善。",
                    }
                ]
            }
        )
        text = "\n".join(bullets)
        self.assertIn("最在意", text)
        self.assertIn("典型动作", text)
        self.assertIn("最反感的误判", text)

    def test_metric_interpretation_section_uses_metric_cards(self) -> None:
        section = _metric_interpretation_section(
            {
                "metric_cards": [
                    {
                        "metric": "收入",
                        "role": "核心结果指标",
                        "business_meaning": "收入代表当前经营结果规模。",
                        "management_impact": "它会影响管理层对规模和增长的判断。",
                        "caution": "需要先确认是否为含税口径。",
                    }
                ]
            },
            {},
        )
        self.assertIsNotNone(section)
        self.assertEqual(section["id"], "metric_interpretation")
        self.assertEqual(section["tables"][0]["title"], "指标解释卡")


if __name__ == "__main__":
    unittest.main()
